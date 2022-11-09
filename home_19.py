# Прочитать сохранённый csv-файл из задания №18 и сохранить данные
# в excel-файл, кроме возраста – столбец с этими данными не нужен.

import csv
import openpyxl



with open('task_02.csv', 'r', encoding='utf-8') as f:
    output_data = csv.reader(f)


Date_of_Birth = ['19990322', '19970719', '19851101', '19700229', '20010517', '19960803', 'Telephone']
Name = ['Tom', 'Alex', 'Nick', 'Sam', 'Sergey', 'David']


wb = openpyxl.Workbook()
print(wb.sheetnames)

wb.create_sheet(title='Первый лист', index=0)
print(wb.sheetnames)

wb.remove(wb['Sheet'])
print(wb.sheetnames)

sheet = wb['Первый лист']
print(sheet)

for row_index, row in enumerate((Date_of_Birth, Name)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index+1, column=col_index+1)
        cell.value = value

wb.save('task_03.xlsx')
